import os
import pandas as pd
import json

UPLOAD_FOLDER = "uploads"

def save_file(file):
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)

    filepath = os.path.join(UPLOAD_FOLDER, file.filename)
    file.save(filepath)
    return filepath


def load_csv(filepath):
    """
    Robust CSV loader with multi-level fallback strategy
    Handles: inconsistent columns, different delimiters, malformed rows, encoding issues
    """
    attempts = [
        # Attempt 1: Standard CSV reading (default)
        {'kwargs': {}, 'desc': 'Standard CSV'},
        # Attempt 2: Skip bad lines, use Python engine (handles inconsistent columns)
        {'kwargs': {'on_bad_lines': 'skip', 'engine': 'python'}, 'desc': 'Skip bad lines'},
        # Attempt 3: Different delimiter (semicolon)
        {'kwargs': {'delimiter': ';', 'on_bad_lines': 'skip', 'engine': 'python'}, 'desc': 'Semicolon delimiter'},
        # Attempt 4: Tab-separated
        {'kwargs': {'delimiter': '\t', 'on_bad_lines': 'skip', 'engine': 'python'}, 'desc': 'Tab-separated'},
        # Attempt 5: Auto-detect delimiter
        {'kwargs': {'sep': None, 'engine': 'python', 'on_bad_lines': 'skip'}, 'desc': 'Auto-detect delimiter'},
        # Attempt 6: Handle encoding issues
        {'kwargs': {'encoding': 'latin-1', 'on_bad_lines': 'skip', 'engine': 'python'}, 'desc': 'Latin-1 encoding'},
        # Attempt 7: Very lenient parsing
        {'kwargs': {'encoding': 'utf-8', 'on_bad_lines': 'skip', 'engine': 'python', 'dtype': str}, 'desc': 'Lenient parsing'},
    ]
    
    for attempt in attempts:
        try:
            df = pd.read_csv(filepath, **attempt['kwargs'])
            if len(df) > 0:  # Ensure we got data
                return df
        except Exception as e:
            continue
    
    raise ValueError(f"Failed to load CSV with all attempted strategies")


def load_excel(filepath):
    """
    Robust Excel loader with fallback options
    Handles: multiple sheets, different sheet names, encoding issues
    Note: If xlrd causes syntax errors, converts Excel to CSV on the fly
    """
    try:
        # Try openpyxl (safest for modern Excel files)
        df = pd.read_excel(filepath, sheet_name=0, engine='openpyxl')
        if len(df) > 0:
            return df
    except SyntaxError:
        # xlrd has Python 2 syntax - convert to CSV using openpyxl directly
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Extract data from worksheet
            data = []
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            
            if data:
                # First row is headers
                df = pd.DataFrame(data[1:], columns=data[0])
                return df
        except Exception:
            pass
    except Exception as e:
        pass
    
    try:
        # Try loading all sheets and concatenate
        all_sheets = pd.read_excel(filepath, sheet_name=None, engine='openpyxl')
        if isinstance(all_sheets, dict):
            dfs = [df for df in all_sheets.values() if len(df) > 0]
            if dfs:
                return pd.concat(dfs, ignore_index=True)
    except Exception:
        pass
    
    raise ValueError(f"Failed to load Excel file")


def load_json(filepath):
    """
    Robust JSON loader with fallback strategies
    Handles: nested JSON, list of objects, JSONL format, different encoding
    """
    file_ext = os.path.splitext(filepath)[1].lower()
    
    # For JSONL files, always try lines=True first
    if file_ext == '.jsonl':
        try:
            df = pd.read_json(filepath, lines=True)
            if len(df) > 0:
                return df
        except Exception:
            pass
    
    # Standard JSON attempts
    attempts = [
        # Attempt 1: Standard JSON-to-dataframe conversion
        {'kwargs': {}, 'desc': 'Standard JSON'},
        # Attempt 2: JSON with lines format (each line is a JSON object)
        {'kwargs': {'lines': True}, 'desc': 'JSON lines'},
        # Attempt 3: JSON with records orient
        {'kwargs': {'orient': 'records'}, 'desc': 'Records orient'},
    ]
    
    for attempt in attempts:
        try:
            df = pd.read_json(filepath, **attempt['kwargs'])
            if len(df) > 0:
                return df
        except Exception:
            pass
    
    # Last attempt: Manual JSON parsing
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            # Try to detect if it's JSONL
            content = f.read()
        
        # Try JSONL format (one JSON per line)
        lines = content.strip().split('\n')
        records = []
        for line in lines:
            if line.strip():
                try:
                    records.append(json.loads(line))
                except:
                    pass
        
        if records and all(isinstance(r, dict) for r in records):
            return pd.DataFrame(records)
        
        # Try standard JSON
        data = json.loads(content)
        
        # Handle various JSON structures
        if isinstance(data, list):
            # List of objects
            if all(isinstance(item, dict) for item in data):
                return pd.DataFrame(data)
        elif isinstance(data, dict):
            # Single object or nested structure
            # Try to find array-like structures
            for key, value in data.items():
                if isinstance(value, list) and value and isinstance(value[0], dict):
                    return pd.DataFrame(value)
            # If no array found, wrap the dict
            return pd.DataFrame([data])
    except Exception:
        pass
    
    raise ValueError(f"Failed to load JSON file with all attempted strategies")


def load_file(filepath):
    """
    Universal file loader - detects format and uses appropriate robust loader
    Supports: CSV, Excel (.xlsx, .xls), JSON, JSONL
    """
    file_ext = os.path.splitext(filepath)[1].lower()
    
    if file_ext == '.csv':
        return load_csv(filepath)
    
    elif file_ext in ['.xlsx', '.xls']:
        return load_excel(filepath)
    
    elif file_ext in ['.json', '.jsonl']:
        return load_json(filepath)
    
    else:
        raise ValueError(f"Unsupported file type: {file_ext}. Supported: CSV, JSON, JSONL, XLS, XLSX")


# Backward compatibility
def load_csv_legacy(filepath):
    return load_file(filepath)